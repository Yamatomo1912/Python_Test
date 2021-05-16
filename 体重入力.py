import datetime
import os
import openpyxl

def isfloat(paramter):
    if not paramter.isdecimal():
        try:
            float(paramter)
            return True
        except ValueError:
            return False
    else:
        return False

def Input_data(str_message):
    def DecimalJudg(data):
        if isfloat(data) == False:
            print("数値以外の文字を入力しないでください\nもう一度入力してください。\n")
            return False
        return True

    input_judg = True
    while (input_judg == True):
        data = input(str_message)
        if DecimalJudg(data) == False:
            input_judg = True
            continue
        input_judg = False
    return data



if __name__ == '__main__':
    
    print("20時になりました、体重計に乗って体重・体脂肪率等記録してください。")
    message_list = [
        "体重(kg)を入力してください。\n",
        "BMI値を入力してください。\n",
        "体脂肪率(%)を入力してください。\n",
        "体水分率(%)を入力してください。\n",
        "骨格筋率(%)を入力してください。\n",
        "基礎代謝(kcal)を入力してください。\n",
        "除脂肪体重(kg)を入力してください。\n",
        "皮下脂肪(%)を入力してください。\n",
        "内臓脂肪(%)を入力してください。\n",
        "筋肉の重さ(kg)を入力してください。\n",
        "骨量(kg)を入力してください。\n",
        "タンパク質(%)を入力してください。\n",
        "体内年齢を入力してください。\n"]
    

    data = []
    for message in message_list:
        data.append(Input_data(message))

    # データ作成
    koumoku_list = []
    for m in message_list:
        koumoku_list.append(m.replace("を入力してください。\n",""))

    now = datetime.datetime.now()
    filename =  str(now.year)+"年.xlsx"
    

    if os.path.exists(filename) == False: # ファイルが存在しなかった場合(新規作成)
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0] # 先頭のシート取得
        ws.title = str(now.month).lstrip("0")+"月" # シート名変更

        # 項目名追加
        count = 2
        ws.cell(1,1).value = "日付"
        for k in koumoku_list:
            ws.cell(1,count).value = k
            count += 1
        
        # 新規データ追加
        count = 2
        ws.cell(2,1).value = str(now.month).lstrip("0")+"月"+str(now.day).lstrip("0")+"日"
        for d in data:
            ws.cell(2,count).value = float(d)
            count += 1
        wb.save(filename)
    else: # 既存データが存在した場合
        workbook = openpyxl.load_workbook(filename)

        # シート名チェック
        worksheet = None
        sheet_check = False # Trueの場合はシートが存在する
        for ws in workbook.sheetnames:
            if ws == str(now.month).lstrip("0")+"月":
                worksheet = workbook[ws] # シートが存在する場合そのシートを選択する
                sheet_check = True 
        
        if sheet_check == False: # シートが存在しない場合は新規で追加する
            worksheet = workbook.create_sheet(title=str(now.month).lstrip("0")+"月")
            # 項目名追加
            count = 2
            worksheet.cell(1,1).value = "日付"
            for k in koumoku_list:
                worksheet.cell(1,count).value = k
                count += 1 


        # データ追加
        rows = worksheet.max_row+1
        worksheet.cell(rows,1).value = str(now.month).lstrip("0")+"月"+str(now.day).lstrip("0")+"月"
        count = 2
        for d in data:
            worksheet.cell(rows,count).value = float(d)
            count += 1
        workbook.save(filename)